# blueprints/kpi.py
from datetime import datetime, date, time, timedelta
from io import BytesIO

from flask import Blueprint, render_template, request, send_file
from flask_login import login_required, current_user
from sqlalchemy import func

from models import db, Ticket, Building, MAINT_DEPTS

bp = Blueprint("kpi", __name__)

# ✅ SLA hours (as provided)
# Urgent: 6h, High: 16h, Medium: 24h, Low: 48h
# + compatibility: emergency treated like urgent
SLA_HOURS = {
    "urgent": 6,
    "emergency": 6,
    "high": 16,
    "medium": 24,
    "low": 48,
}


def _parse_date(value: str):
    try:
        return date.fromisoformat(value)
    except Exception:
        return None


def _dt_range(d_from: date, d_to: date):
    start_dt = datetime.combine(d_from, time.min)
    end_dt = datetime.combine(d_to, time.max)
    return start_dt, end_dt


def _normalize_priority(p: str) -> str:
    if not p:
        return ""
    s = str(p).strip().lower()
    if "urgent" in s:
        return "urgent"
    if "emergency" in s:
        return "emergency"
    if "high" in s:
        return "high"
    if "medium" in s:
        return "medium"
    if "low" in s:
        return "low"
    return s


def _get_sla_hours(priority: str):
    key = _normalize_priority(priority)
    return SLA_HOURS.get(key)


def _normalize_dept(d: str) -> str:
    if not d:
        return ""
    return str(d).strip().lower()


def _dept_scope(base_query, dept: str):
    if dept:
        return base_query.filter(Ticket.maintenance_dept == dept)
    return base_query


def _aging_buckets_for_open(open_now_q, now_utc: datetime):
    bucket_0_24 = bucket_1_3 = bucket_4_7 = bucket_7_plus = 0

    for (c_at,) in open_now_q.with_entities(Ticket.created_at).all():
        if not c_at:
            continue
        hours = (now_utc - c_at).total_seconds() / 3600.0
        if hours <= 24:
            bucket_0_24 += 1
        elif hours <= 72:
            bucket_1_3 += 1
        elif hours <= 168:
            bucket_4_7 += 1
        else:
            bucket_7_plus += 1

    return [
        ("0-24 hours", bucket_0_24),
        ("1-3 days", bucket_1_3),
        ("4-7 days", bucket_4_7),
        ("> 7 days", bucket_7_plus),
    ], bucket_7_plus


def _sla_eval_for_query(created_q):
    """
    SLA Compliance for tickets CREATED in range.
    Returns:
      met, breached, pending, unknown, rate, per_priority_rows
    """
    sla_met = sla_breached = sla_pending = sla_unknown = 0

    per = {
        "urgent": {"met": 0, "breached": 0, "pending": 0, "unknown": 0},
        "emergency": {"met": 0, "breached": 0, "pending": 0, "unknown": 0},
        "high": {"met": 0, "breached": 0, "pending": 0, "unknown": 0},
        "medium": {"met": 0, "breached": 0, "pending": 0, "unknown": 0},
        "low": {"met": 0, "breached": 0, "pending": 0, "unknown": 0},
    }

    rows = created_q.with_entities(Ticket.priority, Ticket.created_at, Ticket.closed_at, Ticket.status).all()
    for pr, c_at, cl_at, st in rows:
        pr_key = _normalize_priority(pr)
        sla_h = _get_sla_hours(pr)

        if pr_key not in per:
            sla_unknown += 1
            continue

        if sla_h is None or not c_at:
            per[pr_key]["unknown"] += 1
            sla_unknown += 1
            continue

        if (st or "").lower() != "closed" or cl_at is None:
            sla_pending += 1
            per[pr_key]["pending"] += 1
            continue

        duration_h = (cl_at - c_at).total_seconds() / 3600.0
        if duration_h <= sla_h:
            sla_met += 1
            per[pr_key]["met"] += 1
        else:
            sla_breached += 1
            per[pr_key]["breached"] += 1

    scored = sla_met + sla_breached
    rate = round((sla_met / scored) * 100.0, 2) if scored else None

    per_priority_rows = []
    for key in ["urgent", "high", "medium", "low"]:
        met = per[key]["met"]
        br = per[key]["breached"]
        pend = per[key]["pending"]
        unk = per[key]["unknown"]
        scored2 = met + br
        rate2 = round((met / scored2) * 100.0, 2) if scored2 else None
        per_priority_rows.append({
            "priority": key.title(),
            "sla_hours": SLA_HOURS.get(key),
            "met": met,
            "breached": br,
            "pending": pend,
            "unknown": unk,
            "rate": rate2,
        })

    return sla_met, sla_breached, sla_pending, sla_unknown, rate, per_priority_rows


def _compute_kpi_payload(d_from: date, d_to: date, selected_dept: str, dept_locked: bool):
    """
    Central function to compute everything (used by page and excel export).
    """
    start_dt, end_dt = _dt_range(d_from, d_to)

    base_scope = Ticket.query
    scope = _dept_scope(base_scope, selected_dept)

    # KPIs within range
    created_q = scope.filter(Ticket.created_at >= start_dt, Ticket.created_at <= end_dt)
    created_count = created_q.count()

    closed_in_range_q = scope.filter(
        Ticket.status == "closed",
        Ticket.closed_at != None,
        Ticket.closed_at >= start_dt,
        Ticket.closed_at <= end_dt
    )
    closed_count = closed_in_range_q.count()

    open_in_range_count = scope.filter(
        Ticket.status != "closed",
        Ticket.created_at >= start_dt,
        Ticket.created_at <= end_dt
    ).count()

    closed_rows = closed_in_range_q.with_entities(Ticket.created_at, Ticket.closed_at).all()
    durations_hours = []
    for c_at, cl_at in closed_rows:
        if c_at and cl_at:
            durations_hours.append((cl_at - c_at).total_seconds() / 3600.0)
    avg_close_hours = round(sum(durations_hours) / len(durations_hours), 2) if durations_hours else None

    # Backlog + aging
    now_utc = datetime.utcnow()
    open_now_q = scope.filter(Ticket.status != "closed")
    backlog_now = open_now_q.count()
    aging_buckets, aging_7_plus = _aging_buckets_for_open(open_now_q, now_utc)

    # SLA
    sla_met, sla_breached, sla_pending, sla_unknown, sla_rate, sla_priority_rows = _sla_eval_for_query(created_q)

    # Breakdowns
    by_status = (
        created_q.with_entities(Ticket.status, func.count(Ticket.id))
                .group_by(Ticket.status)
                .order_by(func.count(Ticket.id).desc())
                .all()
    )

    by_priority = (
        created_q.with_entities(Ticket.priority, func.count(Ticket.id))
                .group_by(Ticket.priority)
                .order_by(func.count(Ticket.id).desc())
                .all()
    )

    by_building = (
        created_q.join(Building, Building.id == Ticket.building_id, isouter=True)
                .with_entities(Building.name, func.count(Ticket.id))
                .group_by(Building.name)
                .order_by(func.count(Ticket.id).desc())
                .limit(10)
                .all()
    )

    # Trend (last 31 days max)
    max_days = 31
    trend_end = d_to
    trend_start = max(d_from, d_to - timedelta(days=max_days - 1))
    t_start_dt, t_end_dt = _dt_range(trend_start, trend_end)

    trend_rows = (
        scope.filter(Ticket.created_at >= t_start_dt, Ticket.created_at <= t_end_dt)
             .with_entities(Ticket.created_at)
             .all()
    )
    counts_map = {}
    for (dt,) in trend_rows:
        if not dt:
            continue
        k = dt.date().isoformat()
        counts_map[k] = counts_map.get(k, 0) + 1

    trend_labels, trend_values = [], []
    cur = trend_start
    while cur <= trend_end:
        key = cur.isoformat()
        trend_labels.append(key)
        trend_values.append(counts_map.get(key, 0))
        cur += timedelta(days=1)

    # Dept comparison (only for ALL and not locked)
    available_depts = list(MAINT_DEPTS)
    dept_compare_rows = []
    best_sla = None
    worst_sla = None
    worst_aging = None
    show_dept_compare = (selected_dept == "" and not dept_locked)

    if show_dept_compare:
        for d in available_depts:
            dept_scope = base_scope.filter(Ticket.maintenance_dept == d)

            dept_created_q = dept_scope.filter(Ticket.created_at >= start_dt, Ticket.created_at <= end_dt)
            dept_created = dept_created_q.count()

            dept_closed_q = dept_scope.filter(
                Ticket.status == "closed",
                Ticket.closed_at != None,
                Ticket.closed_at >= start_dt,
                Ticket.closed_at <= end_dt
            )
            dept_closed = dept_closed_q.count()

            dept_open_now_q = dept_scope.filter(Ticket.status != "closed")
            dept_backlog = dept_open_now_q.count()

            _, dept_aging_7_plus = _aging_buckets_for_open(dept_open_now_q, now_utc)

            d_met, d_br, d_pend, d_unk, d_rate, _ = _sla_eval_for_query(dept_created_q)

            row = {
                "dept": d.upper(),
                "created": dept_created,
                "closed": dept_closed,
                "backlog": dept_backlog,
                "sla_met": d_met,
                "sla_breached": d_br,
                "sla_pending": d_pend,
                "sla_rate": d_rate,
                "aging_7_plus": dept_aging_7_plus,
            }
            dept_compare_rows.append(row)

        # ✅ sort SLA rate desc, None last
        dept_compare_rows.sort(key=lambda r: (r["sla_rate"] is None, -(r["sla_rate"] or 0)))

        scored = [r for r in dept_compare_rows if r["sla_rate"] is not None]
        if scored:
            best_sla = max(scored, key=lambda r: r["sla_rate"])
            worst_sla = min(scored, key=lambda r: r["sla_rate"])
        worst_aging = max(dept_compare_rows, key=lambda r: r["aging_7_plus"]) if dept_compare_rows else None

    return {
        "start_dt": start_dt,
        "end_dt": end_dt,
        "available_depts": available_depts,

        "created_count": created_count,
        "closed_count": closed_count,
        "open_in_range_count": open_in_range_count,
        "avg_close_hours": avg_close_hours,

        "backlog_now": backlog_now,
        "aging_buckets": aging_buckets,
        "aging_7_plus": aging_7_plus,

        "sla_met": sla_met,
        "sla_breached": sla_breached,
        "sla_pending": sla_pending,
        "sla_unknown": sla_unknown,
        "sla_rate": sla_rate,
        "sla_priority_rows": sla_priority_rows,

        "by_status": by_status,
        "by_priority": by_priority,
        "by_building": by_building,

        "trend_labels": trend_labels,
        "trend_values": trend_values,

        "show_dept_compare": show_dept_compare,
        "dept_compare_rows": dept_compare_rows,
        "best_sla": best_sla,
        "worst_sla": worst_sla,
        "worst_aging": worst_aging,
    }


def _resolve_filters():
    """
    Resolve date range + dept with role-based locking.
    """
    today = date.today()
    first_day = today.replace(day=1)

    from_str = (request.args.get("from") or "").strip()
    to_str = (request.args.get("to") or "").strip()

    d_from = _parse_date(from_str) or first_day
    d_to = _parse_date(to_str) or today
    if d_to < d_from:
        d_from, d_to = d_to, d_from

    requested_dept = _normalize_dept(request.args.get("dept") or "")
    available_depts = list(MAINT_DEPTS)

    user_forced_dept = ""
    if current_user.role in ("supervisor", "technician") and getattr(current_user, "maintenance_dept", None):
        user_forced_dept = _normalize_dept(current_user.maintenance_dept)

    if user_forced_dept:
        selected_dept = user_forced_dept
        dept_locked = True
    else:
        if requested_dept and requested_dept not in available_depts:
            requested_dept = ""
        selected_dept = requested_dept  # "" means All
        dept_locked = False

    return d_from, d_to, selected_dept, dept_locked, available_depts


@bp.get("/kpi")
@login_required
def kpi():
    d_from, d_to, selected_dept, dept_locked, available_depts = _resolve_filters()
    data = _compute_kpi_payload(d_from, d_to, selected_dept, dept_locked)

    return render_template(
        "kpi.html",
        # filters
        date_from=d_from.isoformat(),
        date_to=d_to.isoformat(),
        maint_depts=available_depts,
        selected_dept=selected_dept,
        dept_locked=dept_locked,

        # stats
        created_count=data["created_count"],
        closed_count=data["closed_count"],
        open_in_range_count=data["open_in_range_count"],
        avg_close_hours=data["avg_close_hours"],

        backlog_now=data["backlog_now"],
        aging_buckets=data["aging_buckets"],
        aging_7_plus=data["aging_7_plus"],

        sla_met=data["sla_met"],
        sla_breached=data["sla_breached"],
        sla_pending=data["sla_pending"],
        sla_unknown=data["sla_unknown"],
        sla_rate=data["sla_rate"],
        sla_priority_rows=data["sla_priority_rows"],

        by_status=data["by_status"],
        by_priority=data["by_priority"],
        by_building=data["by_building"],

        trend_labels=data["trend_labels"],
        trend_values=data["trend_values"],

        show_dept_compare=data["show_dept_compare"],
        dept_compare_rows=data["dept_compare_rows"],
        best_sla=data["best_sla"],
        worst_sla=data["worst_sla"],
        worst_aging=data["worst_aging"],
    )


@bp.get("/kpi/export.xlsx")
@login_required
def kpi_export_xlsx():
    """
    Export the SAME KPI view to Excel (.xlsx) with the same filters.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment
    except Exception:
        return "openpyxl is not installed. Run: pip install openpyxl", 500

    d_from, d_to, selected_dept, dept_locked, available_depts = _resolve_filters()
    data = _compute_kpi_payload(d_from, d_to, selected_dept, dept_locked)

    dept_label = selected_dept.upper() if selected_dept else "ALL"
    filename = f"KPI_{dept_label}_{d_from.isoformat()}_to_{d_to.isoformat()}.xlsx"

    wb = Workbook()

    def _style_header(row_cells):
        for c in row_cells:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

    def _autosize(ws):
        # ✅ FIXED: proper autosize loop
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val = "" if cell.value is None else str(cell.value)
                    if len(val) > max_len:
                        max_len = len(val)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)

    # ---------------- Sheet: Summary ----------------
    ws = wb.active
    ws.title = "Summary"

    ws.append(["Department", dept_label])
    ws.append(["From", d_from.isoformat()])
    ws.append(["To", d_to.isoformat()])
    ws.append([])
    ws.append(["Metric", "Value"])
    _style_header(ws[5])

    ws.append(["Created (in range)", data["created_count"]])
    ws.append(["Open (in range)", data["open_in_range_count"]])
    ws.append(["Closed (in range)", data["closed_count"]])
    ws.append(["Avg Close Time (hours)", data["avg_close_hours"] if data["avg_close_hours"] is not None else "-"])
    ws.append([])
    ws.append(["Backlog (Open Now)", data["backlog_now"]])
    ws.append(["> 7 days open (Open Now)", data["aging_7_plus"]])
    ws.append([])
    ws.append(["SLA Met", data["sla_met"]])
    ws.append(["SLA Breached", data["sla_breached"]])
    ws.append(["SLA Pending", data["sla_pending"]])
    ws.append(["SLA Rate (%)", data["sla_rate"] if data["sla_rate"] is not None else "-"])

    _autosize(ws)

    # ---------------- Sheet: SLA ----------------
    ws = wb.create_sheet("SLA")
    ws.append(["Priority", "SLA (hours)", "Met", "Breached", "Pending", "Rate (%)"])
    _style_header(ws[1])

    for r in data["sla_priority_rows"]:
        ws.append([
            r["priority"],
            r["sla_hours"],
            r["met"],
            r["breached"],
            r["pending"],
            r["rate"] if r["rate"] is not None else "-"
        ])
    _autosize(ws)

    # ---------------- Sheet: Aging ----------------
    ws = wb.create_sheet("Aging")
    ws.append(["Bucket", "Count (Open Now)"])
    _style_header(ws[1])
    for label, cnt in data["aging_buckets"]:
        ws.append([label, cnt])
    _autosize(ws)

    # ---------------- Sheet: Trend ----------------
    ws = wb.create_sheet("Trend")
    ws.append(["Date", "Created Count"])
    _style_header(ws[1])
    for d, c in zip(data["trend_labels"], data["trend_values"]):
        ws.append([d, c])
    _autosize(ws)

    # ---------------- Sheet: Top Buildings ----------------
    ws = wb.create_sheet("Top Buildings")
    ws.append(["Building", "Count (Created in Range)"])
    _style_header(ws[1])
    for bname, cnt in data["by_building"]:
        ws.append([bname if bname else "-", cnt])
    _autosize(ws)

    # ---------------- Sheet: Dept Compare (only if ALL) ----------------
    if data["show_dept_compare"]:
        ws = wb.create_sheet("Dept Compare")
        ws.append(["Dept", "Created", "Closed", "Backlog", "SLA Met", "SLA Breached", "SLA Pending", "SLA Rate (%)", "> 7 days open"])
        _style_header(ws[1])
        for r in data["dept_compare_rows"]:
            ws.append([
                r["dept"],
                r["created"],
                r["closed"],
                r["backlog"],
                r["sla_met"],
                r["sla_breached"],
                r["sla_pending"],
                r["sla_rate"] if r["sla_rate"] is not None else "-",
                r["aging_7_plus"],
            ])
        _autosize(ws)

    # Save to memory
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
